import os
from io import BytesIO

import django_tables2 as tables
import openpyxl
from django.http import HttpResponse
from django.utils.html import format_html
from django_tables2.export import TableExport
from openpyxl.styles import Font

from Dj_TPS.settings import BASE_DIR
from .models import *


class SalesReportTable(tables.Table):
    class Meta:
        model = Sales
        template_name = "django_tables2/bootstrap5.html"
        fields = (
            "date", "store", "staff", "photographer", "sale_type", "payment_type", "photo_count", "sum", "cl_email", "cl_phone", "comment"
        )
        attrs = {
            "class": "table table-striped table-hover",
            "thead": {"class": "table-light"}
        }
        order_by = "-date"


class SalaryWeekReportTable(tables.Table):
    week = tables.TemplateColumn(
        template_code='{{ record.week_begin|date:"d F" }} - {{ record.week_end|date:"d F (Y)" }}',
        verbose_name="Неделя",
        order_by=('week_begin',)  # Сортировка по началу недели
    )

    class Meta:
        model = SalaryWeekly
        template_name = "django_tables2/bootstrap5.html"
        fields = ("week", "staff", "cash_box_week", "salary_sum", "cash_withdrawn", "to_pay", "paid_out")

        attrs = {
            "class": "table table-striped table-hover",
            "thead": {"class": "table-light"}
        }
        order_by = "-week_begin"


class SalaryReportTable(tables.Table):
    class Meta:
        model = Salary
        template_name = "django_tables2/bootstrap5.html"
        fields = ("store", "staff", "date", "cash_box", "salary_sum", "cnt_logs")

        attrs = {
            "class": "table table-striped table-hover",
            "thead": {"class": "table-light"}
        }
        order_by = "-date"


def export_to_excel_template(table, queryset, date_from, date_by, model_name):
    """Экспорт данных в существующий Excel шаблон"""

    template_path = os.path.join(BASE_DIR, 'tph_system', 'static', 'tph_system', 'files', 'excel_template_Sales.xlsx')
    if not os.path.exists(template_path):
        print(f"Файл шаблона не найден: {template_path}")
        return standard_export(table, date_from, date_by, 'xlsx', model_name)

    try:
        # Загружаем шаблон
        workbook = openpyxl.load_workbook(template_path)

        # Получаем или создаем первый лист
        if 'Данные' in workbook.sheetnames:
            worksheet = workbook['Данные']
            # Очищаем существующие данные (оставляя заголовки если нужно)
            worksheet.delete_rows(2, worksheet.max_row)  # удаляем все строки кроме заголовка
        else:
            # Если листа 'Данные' нет - используем первый лист
            worksheet = workbook.active
            worksheet.title = 'Данные'
            # Очищаем все данные
            worksheet.delete_rows(1, worksheet.max_row)

        # Записываем заголовки
        headers = [column.verbose_name for column in table.columns]
        for col_idx, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col_idx, value=header)
            worksheet.cell(row=1, column=col_idx).font = Font(bold=True)

        # Записываем данные
        for row_idx, record in enumerate(queryset, 2):  # начинаем с 2 строки
            worksheet.cell(row=row_idx, column=1, value=record.date)
            worksheet.cell(row=row_idx, column=1).number_format = 'dd.mm.yyyy'
            worksheet.cell(row=row_idx, column=2, value=str(record.store))
            worksheet.cell(row=row_idx, column=3, value=str(record.staff))
            worksheet.cell(row=row_idx, column=4, value=str(record.photographer))
            worksheet.cell(row=row_idx, column=5, value=str(record.sale_type))
            worksheet.cell(row=row_idx, column=6, value=str(record.payment_type))
            worksheet.cell(row=row_idx, column=7, value=int(record.photo_count or 0))
            worksheet.cell(row=row_idx, column=8, value=int(record.sum or 0))
            worksheet.cell(row=row_idx, column=9, value=str(record.cl_email))
            worksheet.cell(row=row_idx, column=10, value=str(record.cl_phone))
            worksheet.cell(row=row_idx, column=11, value=str(record.comment))

        # Сохраняем в память
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        # Создаем HTTP response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"{model_name} Report from {date_from or ""} to {date_by or ""}"
        response['Content-Disposition'] = 'attachment; filename="' + filename + '.xlsx"'

        return response

    except Exception as e:
        # В случае ошибки - fallback на стандартный экспорт
        print(e)
        return standard_export(table, date_from, date_by, 'xlsx', model_name)


def standard_export(table, date_from, date_by, export_format, model_name):
    """Стандартный экспорт таблицы"""
    exporter = TableExport(export_format, table)
    return exporter.response(f"{model_name} Report from {date_from or ""} to {date_by or ""}.{export_format}")
